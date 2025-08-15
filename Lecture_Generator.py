from ortools.sat.python import cp_model
import datetime
import random
import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

THEORY_CLASS_DURATION = 1.0
START_TIME = "8:15"

class LectureScheduler:
    def __init__(self, department_data, lab_schedule):
        self.morning_break = "10:15-10:30"
        self.lunch_break = "12:30-1:15"
        self.evening_break = "3:15-3:30"
        self.morning_slots = ["8:15-9:15", "9:15-10:15"]
        self.midday_slots = ["10:30-11:30", "11:30-12:30"]
        self.afternoon_slots = ["1:15-2:15", "2:15-3:15"]
        self.evening_slots = ["3:30-4:30", "4:30-5:30"]
        self.all_time_slots = (
            self.morning_slots + [self.morning_break] +
            self.midday_slots + [self.lunch_break] +
            self.afternoon_slots + [self.evening_break] +
            self.evening_slots
        )
        self.break_slots = [self.morning_break, self.lunch_break, self.evening_break]
        self.years = department_data["years"]
        self.classes_per_year = department_data["classes_per_year"]
        self.all_classes = []
        for year in self.years:
            if year == "Second Year":
                year_prefix = "SE"
            elif year == "Third Year":
                year_prefix = "TE"
            elif year == "Fourth Year":
                year_prefix = "BE"
            else:
                year_prefix = year[:2].upper()
            for i in range(1, self.classes_per_year + 1):
                self.all_classes.append(f"{year_prefix}{i}")
        self.teachers = department_data["teachers"]
        self.rooms = department_data["rooms"]
        self.subjects_by_year = department_data["subjects_by_year"]
        self.course_structure = department_data["course_structure"]
        self.teacher_assignments = department_data["teacher_assignments"]
        self.days = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday"]
        self.teacher_availability = department_data["teacher_availability"]
        self.available_slots = [slot for slot in self.all_time_slots if slot not in self.break_slots]
        self.lab_schedule = lab_schedule  # Lab schedule from generator1
        self.model = cp_model.CpModel()
        self.assignments = {}
        self.class_timetables = {}
        # Define allowed lecture slots for each year
        self.te_lecture_slots = ["8:15-9:15", "9:15-10:15", "10:30-11:30", "11:30-12:30", "1:15-2:15", "2:15-3:15", "3:30-4:30"]
        self.se_lecture_slots = ["8:15-9:15", "9:15-10:15","10:30-11:30", "11:30-12:30", "1:15-2:15", "2:15-3:15", "3:30-4:30", "4:30-5:30"]

    def _get_year_from_class(self, class_name):
        prefix = class_name[:2].upper()
        if prefix == "SE":
            return "Second Year"
        elif prefix == "TE":
            return "Third Year"
        elif prefix == "BE":
            return "Fourth Year"
        else:
            print(f"Warning: Unknown prefix '{prefix}' in class name '{class_name}'. Defaulting to Second Year.")
            return "Second Year"

    def get_lab_slots_for_class(self, class_name):
        """Get all time slots where labs are scheduled for a class"""
        lab_slots = set()
        if class_name in self.lab_schedule:
            for day in self.days:
                for time_slot in self.available_slots:
                    labs = self.lab_schedule[class_name]["labs"][day].get(time_slot, {})
                    for batch, lab in labs.items():
                        if lab and not lab.get("continued", False):
                            # Add both consecutive slots for the lab
                            lab_slots.add((day, time_slot))
                            if time_slot in ["8:15-9:15", "10:30-11:30", "1:15-2:15", "3:30-4:30"]:
                                # Add the second hour slot
                                if time_slot == "8:15-9:15":
                                    lab_slots.add((day, "9:15-10:15"))
                                elif time_slot == "10:30-11:30":
                                    lab_slots.add((day, "11:30-12:30"))
                                elif time_slot == "1:15-2:15":
                                    lab_slots.add((day, "2:15-3:15"))
                                elif time_slot == "3:30-4:30":
                                    lab_slots.add((day, "4:30-5:30"))
        return lab_slots

    def create_lecture_variables(self):
        """Create variables for lecture scheduling"""
        for class_name in self.all_classes:
            year = self._get_year_from_class(class_name)
            if year == "Third Year":
                allowed_slots = self.te_lecture_slots
                allowed_rooms = ["507"]  # TE lectures only in 507
            elif year == "Second Year":
                allowed_slots = self.se_lecture_slots
                allowed_rooms = ["506"]  # SE lectures only in 506
            else:
                allowed_slots = self.available_slots
                allowed_rooms = self.rooms
            subjects = self.subjects_by_year[year]
            for subject in subjects:
                if subject in self.course_structure and self.course_structure[subject].get("lectures", 0) > 0:
                    for lecture_index in range(self.course_structure[subject]["lectures"]):
                        for day in self.days:
                            for time_slot in allowed_slots:
                                teacher = self.teacher_assignments[year][subject]
                                for room in allowed_rooms:  # Use allowed_rooms instead of self.rooms
                                    key = (class_name, subject, "lecture", lecture_index, day, time_slot, teacher, room)
                                    self.assignments[key] = self.model.NewBoolVar(f"{class_name}{subject}_Lecture{lecture_index}{day}{time_slot}{teacher}{room}")

    def add_lecture_constraints(self):
        """Add constraints for lecture scheduling"""
        # Each lecture must be scheduled exactly once for each class
        for class_name in self.all_classes:
            year = self._get_year_from_class(class_name)
            if year == "Third Year":
                allowed_slots = self.te_lecture_slots
                allowed_rooms = ["507"]  # TE lectures only in 507
            elif year == "Second Year":
                allowed_slots = self.se_lecture_slots
                allowed_rooms = ["506"]  # SE lectures only in 506
            else:
                allowed_slots = self.available_slots
                allowed_rooms = self.rooms
            subjects = self.subjects_by_year[year]
            for subject in subjects:
                if subject in self.course_structure and self.course_structure[subject].get("lectures", 0) > 0:
                    for lecture_index in range(self.course_structure[subject]["lectures"]):
                        lecture_vars = []
                        for day in self.days:
                            for time_slot in allowed_slots:
                                teacher = self.teacher_assignments[year][subject]
                                for room in allowed_rooms:  # Use allowed_rooms instead of self.rooms
                                    key = (class_name, subject, "lecture", lecture_index, day, time_slot, teacher, room)
                                    if key in self.assignments:
                                        lecture_vars.append(self.assignments[key])
                        if lecture_vars:
                            self.model.Add(sum(lecture_vars) == 1)

        # Teacher availability constraints
        for teacher in self.teachers:
            for day in self.days:
                for time_slot in self.available_slots:
                    teacher_vars = []
                    for key, var in self.assignments.items():
                        class_name, subject, activity_type, lecture_index, day_key, time_slot_key, teacher_key, room = key
                        if day_key == day and time_slot_key == time_slot and teacher_key == teacher:
                            teacher_vars.append(var)
                    if teacher_vars:
                        if teacher in self.teacher_availability and day in self.teacher_availability[teacher]:
                            if time_slot not in self.teacher_availability[teacher][day]:
                                self.model.Add(sum(teacher_vars) == 0)
                        self.model.Add(sum(teacher_vars) <= 1)
        
        # Prevent teacher conflicts with lab schedule
        for class_name in self.all_classes:
            if class_name in self.lab_schedule:
                for day in self.days:
                    for time_slot in self.available_slots:
                        # Check if any teacher has a lab at this exact time slot
                        labs = self.lab_schedule[class_name]["labs"][day].get(time_slot, {})
                        for batch, lab in labs.items():
                            if lab:  # Check both first hour and continued labs
                                lab_teacher = lab.get("teacher")
                                if lab_teacher:
                                    # Prevent this teacher from being assigned to any lecture at this EXACT time
                                    lecture_vars = []
                                    for key, var in self.assignments.items():
                                        lecture_class, subject, activity_type, lecture_index, day_key, time_slot_key, teacher_key, room = key
                                        if day_key == day and time_slot_key == time_slot and teacher_key == lab_teacher:
                                            lecture_vars.append(var)
                                    if lecture_vars:
                                        # Teacher cannot have any lecture when they have a lab
                                        self.model.Add(sum(lecture_vars) == 0)

        # Room availability constraints
        for room in self.rooms:
            for day in self.days:
                for time_slot in self.available_slots:
                    room_vars = []
                    for key, var in self.assignments.items():
                        class_name, subject, activity_type, lecture_index, day_key, time_slot_key, teacher, room_key = key
                        if day_key == day and time_slot_key == time_slot and room_key == room:
                            room_vars.append(var)
                    if room_vars:
                        self.model.Add(sum(room_vars) <= 1)

        # Class availability constraints (no overlapping lectures for the same class)
        for class_name in self.all_classes:
            year = self._get_year_from_class(class_name)
            if year == "Third Year":
                allowed_slots = self.te_lecture_slots
            elif year == "Second Year":
                allowed_slots = self.se_lecture_slots
            else:
                allowed_slots = self.available_slots
            for day in self.days:
                for time_slot in allowed_slots:
                    class_vars = []
                    for key, var in self.assignments.items():
                        class_name_key, subject, activity_type, lecture_index, day_key, time_slot_key, teacher, room = key
                        if class_name_key == class_name and day_key == day and time_slot_key == time_slot:
                            class_vars.append(var)
                    if class_vars:
                        self.model.Add(sum(class_vars) <= 1)

        # Lab conflict constraints (no lectures when labs are scheduled)
        for class_name in self.all_classes:
            year = self._get_year_from_class(class_name)
            if year == "Third Year":
                allowed_slots = self.te_lecture_slots
            elif year == "Second Year":
                allowed_slots = self.se_lecture_slots
            else:
                allowed_slots = self.available_slots
            lab_slots = self.get_lab_slots_for_class(class_name)
            for day, time_slot in lab_slots:
                if time_slot not in allowed_slots:
                    continue
                class_vars = []
                for key, var in self.assignments.items():
                    class_name_key, subject, activity_type, lecture_index, day_key, time_slot_key, teacher, room = key
                    if class_name_key == class_name and day_key == day and time_slot_key == time_slot:
                        class_vars.append(var)
                if class_vars:
                    self.model.Add(sum(class_vars) == 0)

        # No more than one lecture per subject per day per class
        for class_name in self.all_classes:
            year = self._get_year_from_class(class_name)
            if year == "Third Year":
                allowed_slots = self.te_lecture_slots
                allowed_rooms = ["507"]  # TE lectures only in 507
            elif year == "Second Year":
                allowed_slots = self.se_lecture_slots
                allowed_rooms = ["506"]  # SE lectures only in 506
            else:
                allowed_slots = self.available_slots
                allowed_rooms = self.rooms
            subjects = self.subjects_by_year[year]
            for subject in subjects:
                if subject in self.course_structure and self.course_structure[subject].get("lectures", 0) > 0:
                    for day in self.days:
                        subject_lecture_vars = []
                        for lecture_index in range(self.course_structure[subject]["lectures"]):
                            for time_slot in allowed_slots:
                                teacher = self.teacher_assignments[year][subject]
                                for room in allowed_rooms:  # Use allowed_rooms instead of self.rooms
                                    key = (class_name, subject, "lecture", lecture_index, day, time_slot, teacher, room)
                                    if key in self.assignments:
                                        subject_lecture_vars.append(self.assignments[key])
                        if subject_lecture_vars:
                            self.model.Add(sum(subject_lecture_vars) <= 1)

    def add_optimization_objective(self):
        """Add optimization objectives to minimize gaps and optimize scheduling"""
        gap_vars = []
        late_session_vars = []
        continuity_vars = []
        slot_index = {slot: idx for idx, slot in enumerate(self.available_slots)}
        num_slots = len(self.available_slots)

        for class_name in self.all_classes:
            for day in self.days:
                # Build combined schedule: lectures + labs for this class on this day
                combined_slot_vars = []
                
                for slot in self.available_slots:
                    slot_has_activity = self.model.NewBoolVar(f"activity_{class_name}_{day}_{slot}")
                    activity_vars = []
                    
                    # Check for lectures
                    for key, var in self.assignments.items():
                        if key[0] == class_name and key[2] == "lecture" and key[4] == day and key[5] == slot:
                            activity_vars.append(var)
                    
                    # Check for labs from the pre-scheduled lab timetable
                    if class_name in self.lab_schedule:
                        labs = self.lab_schedule[class_name]["labs"][day].get(slot, {})
                        for batch, lab in labs.items():
                            if lab and not lab.get("continued", False):
                                # This slot has a lab - add a constant true variable
                                lab_indicator = self.model.NewBoolVar(f"lab_{class_name}_{day}_{slot}_{batch}")
                                self.model.Add(lab_indicator == 1)  # Always true since lab is scheduled
                                activity_vars.append(lab_indicator)
                    
                    # slot_has_activity is true if any lecture or lab is scheduled
                    if activity_vars:
                        self.model.AddBoolOr(activity_vars).OnlyEnforceIf(slot_has_activity)
                        self.model.AddBoolAnd([var.Not() for var in activity_vars]).OnlyEnforceIf(slot_has_activity.Not())
                    else:
                        self.model.Add(slot_has_activity == 0)
                    
                    combined_slot_vars.append(slot_has_activity)

                # Enhanced gap detection for combined schedule
                for i in range(1, num_slots-1):
                    prev_var = combined_slot_vars[i-1]
                    curr_var = combined_slot_vars[i]
                    next_var = combined_slot_vars[i+1]
                    
                    # Gap penalty: previous and next slots have activities, but current doesn't
                    gap = self.model.NewBoolVar(f"gap_{class_name}_{day}_{i}")
                    self.model.AddBoolAnd([prev_var, next_var, curr_var.Not()]).OnlyEnforceIf(gap)
                    self.model.AddBoolOr([prev_var.Not(), next_var.Not(), curr_var]).OnlyEnforceIf(gap.Not())
                    gap_vars.append(gap)

                # Simplified post-break continuity optimization
                post_break_idle_vars = []
                # Encourage activities right after breaks
                for i, slot in enumerate(self.available_slots):
                    if slot in ["10:30-11:30", "1:15-2:15"]:  # First slots after breaks
                        # Create a soft penalty for being idle right after breaks
                        if i < len(combined_slot_vars):
                            post_break_idle = self.model.NewBoolVar(f"post_break_idle_{class_name}_{day}_{i}")
                            # If this slot is empty, it's a post-break idle (soft penalty only)
                            self.model.Add(post_break_idle == combined_slot_vars[i].Not())
                            post_break_idle_vars.append(post_break_idle)
                
                # Simplified compact scheduling: minimize idle slots between activities
                # Count total activities for this class on this day
                total_activities = self.model.NewIntVar(0, num_slots, f"total_activities_{class_name}_{day}")
                self.model.Add(total_activities == sum(combined_slot_vars))
                
                # Simple spread penalty: penalize activities that are far apart
                spread_penalty_vars = []
                for i in range(num_slots):
                    for j in range(i+4, min(i+7, num_slots)):  # Look for activities 4-6 slots apart (reduce penalty range)
                        both_active = self.model.NewBoolVar(f"spread_{class_name}_{day}_{i}_{j}")
                        self.model.AddBoolAnd([combined_slot_vars[i], combined_slot_vars[j]]).OnlyEnforceIf(both_active)
                        self.model.AddBoolOr([combined_slot_vars[i].Not(), combined_slot_vars[j].Not()]).OnlyEnforceIf(both_active.Not())
                        # Add a light penalty for distance
                        spread_penalty_vars.append(both_active)

                # Penalize late sessions for lectures only
                lecture_slot_vars = []
                for slot in self.available_slots:
                    found = False
                    for key, var in self.assignments.items():
                        if key[0] == class_name and key[2] == "lecture" and key[4] == day and key[5] == slot:
                            lecture_slot_vars.append(var)
                            found = True
                            break
                    if not found:
                        lecture_slot_vars.append(None)

                for i, var in enumerate(lecture_slot_vars):
                    if var is not None:
                        weight = i * 2
                        for _ in range(weight):
                            late_session_vars.append(var)

        # Enhanced objective: prioritize continuity and minimize gaps
        self.model.Minimize(
            50 * sum(gap_vars) +        # Very high penalty for gaps in combined schedule
            100 * sum(post_break_idle_vars) + # CRITICAL: Extra high penalty for post-break idle periods
            20 * sum(spread_penalty_vars) + # Penalty for long spans between first and last activity
            2 * sum(late_session_vars)   # Medium penalty for late lecture sessions
        )

    def solve(self, timeout_seconds=120):
        """Solve the lecture scheduling problem"""
        self.create_lecture_variables()
        self.add_lecture_constraints()
        self.add_optimization_objective()
        
        solver = cp_model.CpSolver()
        solver.parameters.max_time_in_seconds = timeout_seconds
        solver.parameters.num_search_workers = 16
        solver.parameters.log_search_progress = True
        
        status = solver.Solve(self.model)
        if status in [cp_model.OPTIMAL, cp_model.FEASIBLE]:
            self._extract_timetables(solver)
            return True
        else:
            print(f"No solution found. Status: {status}")
            return False

    def _extract_timetables(self, solver):
        """Extract the timetable from the solver solution"""
        for class_name in self.all_classes:
            self.class_timetables[class_name] = {"lectures": {}}
            for day in self.days:
                self.class_timetables[class_name]["lectures"][day] = {}
                for time_slot in self.all_time_slots:
                    self.class_timetables[class_name]["lectures"][day][time_slot] = None
                    if time_slot in self.break_slots:
                        if time_slot == self.lunch_break:
                            self.class_timetables[class_name]["lectures"][day][time_slot] = {"subject": "LUNCH BREAK"}
                        elif time_slot == self.morning_break:
                            self.class_timetables[class_name]["lectures"][day][time_slot] = {"subject": "MORNING BREAK"}
                        elif time_slot == self.evening_break:
                            self.class_timetables[class_name]["lectures"][day][time_slot] = {"subject": "EVENING BREAK"}

        for key, var in self.assignments.items():
            if solver.BooleanValue(var):
                class_name, subject, activity_type, lecture_index, day, time_slot, teacher, room = key
                self.class_timetables[class_name]["lectures"][day][time_slot] = {
                    "subject": subject,
                    "teacher": teacher,
                    "room": room,
                    "lecture_index": lecture_index
                }

    def export_combined_timetable_to_excel(self):
        """Export combined timetables to a beautifully formatted Excel file"""
        timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"combined_timetables_{timestamp}.xlsx"
        
        with pd.ExcelWriter(filename, engine='openpyxl') as writer:
            for class_name in self.all_classes:
                if class_name not in self.class_timetables:
                    continue
                
                timetable = self.class_timetables[class_name]
                lab_timetable = self.lab_schedule[class_name] if class_name in self.lab_schedule else None
                
                # Create DataFrame for the timetable
                data = []
                for time_slot in self.all_time_slots:
                    row = [time_slot]
                    for day in self.days:
                        if time_slot in self.break_slots:
                            if time_slot == self.lunch_break:
                                row.append("LUNCH BREAK")
                            elif time_slot == self.morning_break:
                                row.append("MORNING BREAK")
                            elif time_slot == self.evening_break:
                                row.append("EVENING BREAK")
                        else:
                            # Check for lecture
                            lecture = timetable["lectures"][day].get(time_slot)
                            # Check for labs
                            labs = []
                            if lab_timetable:
                                for batch, lab in lab_timetable["labs"][day].get(time_slot, {}).items():
                                    if lab and not lab.get("continued", False):
                                        labs.append(f"{lab['subject']},{batch},{lab['lab']}")
                                    elif lab and lab.get("continued", False):
                                        # Show continued labs in the second hour
                                        labs.append(f"{lab['subject']},{batch},{lab['lab']} (cont.)")
                            # Combine lecture and labs info
                            if lecture:
                                cell = f"{lecture['subject']} ({lecture['teacher']}) ({lecture['room']})"
                                row.append(cell)
                            elif labs:
                                # Format labs with teacher names
                                formatted_labs = []
                                for lab_info in labs:
                                    parts = lab_info.split(',')
                                    if len(parts) >= 3:
                                        subject = parts[0]
                                        batch = parts[1]
                                        lab = parts[2]
                                        year = self._get_year_from_class(class_name)
                                        teacher = self.teacher_assignments[year].get(subject, "Unknown")
                                        formatted_labs.append(f"{subject} ({teacher}) {batch} {lab}")
                                    else:
                                        formatted_labs.append(lab_info)
                                cell = " && ".join(formatted_labs)
                                row.append(cell)
                            else:
                                row.append("---")
                    data.append(row)
                
                df = pd.DataFrame(data, columns=["Time Slot"] + self.days)
                df.to_excel(writer, sheet_name=class_name, index=False)
                
                # Get the workbook and worksheet for styling
                workbook = writer.book
                worksheet = writer.sheets[class_name]
                
                # Apply styling
                self._style_timetable_worksheet(worksheet, class_name)
        
        print(f"\n✅ Combined timetables exported to: {filename}")
        return filename
    
    def _style_timetable_worksheet(self, worksheet, class_name):
        """Apply beautiful styling to the timetable worksheet"""
        # Define colors
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        break_fill = PatternFill(start_color="FFE699", end_color="FFE699", fill_type="solid")
        lab_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        lecture_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
        empty_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        
        # Define fonts
        header_font = Font(color="FFFFFF", bold=True, size=12)
        break_font = Font(bold=True, size=11)
        content_font = Font(size=10)
        
        # Define borders
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Style header row
        for col in range(1, len(self.days) + 2):  # +2 for Time Slot column
            cell = worksheet.cell(row=1, column=col)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = thin_border
        
        # Style data rows
        for row in range(2, worksheet.max_row + 1):
            time_slot = worksheet.cell(row=row, column=1).value
            
            # Style time slot column
            time_cell = worksheet.cell(row=row, column=1)
            time_cell.font = Font(bold=True, size=10)
            time_cell.alignment = Alignment(horizontal='left', vertical='center')
            time_cell.border = thin_border
            
            # Style day columns
            for col in range(2, len(self.days) + 2):
                cell = worksheet.cell(row=row, column=col)
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                cell.border = thin_border
                
                # Apply color coding based on content
                cell_value = cell.value
                if cell_value in ["LUNCH BREAK", "MORNING BREAK", "EVENING BREAK"]:
                    cell.fill = break_fill
                    cell.font = break_font
                elif cell_value and cell_value != "---":
                    if "Lab-" in str(cell_value):  # Lab session
                        cell.fill = lab_fill
                        cell.font = content_font
                    elif "LR-" in str(cell_value):  # Lecture session
                        cell.fill = lecture_fill
                        cell.font = content_font
                else:  # Empty slot
                    cell.fill = empty_fill
                    cell.font = content_font
        
        # Auto-adjust column widths
        for column in worksheet.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 25)  # Cap at 25 characters
            worksheet.column_dimensions[column_letter].width = adjusted_width
        
        # Set row heights
        for row in range(1, worksheet.max_row + 1):
            worksheet.row_dimensions[row].height = 25
        
        # Add title
        title = f"COMBINED TIMETABLE - {class_name} ({self._get_year_from_class(class_name)})"
        worksheet.insert_rows(1)
        title_cell = worksheet.cell(row=1, column=1)
        title_cell.value = title
        title_cell.font = Font(bold=True, size=14, color="366092")
        title_cell.alignment = Alignment(horizontal='center')
        worksheet.merge_cells('A1:G1')  # Merge cells for title
        
        # Add timestamp
        timestamp = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        timestamp_cell = worksheet.cell(row=2, column=1)
        timestamp_cell.value = f"Generated on: {timestamp}"
        timestamp_cell.font = Font(italic=True, size=9, color="666666")
        timestamp_cell.alignment = Alignment(horizontal='center')
        worksheet.merge_cells('A2:G2')  # Merge cells for timestamp
        
        # Adjust data rows after inserting title
        for row in range(4, worksheet.max_row + 1):  # Start from row 4 (after title and timestamp)
            time_slot = worksheet.cell(row=row, column=1).value
            
            # Style time slot column
            time_cell = worksheet.cell(row=row, column=1)
            time_cell.font = Font(bold=True, size=10)
            time_cell.alignment = Alignment(horizontal='left', vertical='center')
            time_cell.border = thin_border
            
            # Style day columns
            for col in range(2, len(self.days) + 2):
                cell = worksheet.cell(row=row, column=col)
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                cell.border = thin_border
                
                # Apply color coding based on content
                cell_value = cell.value
                if cell_value in ["LUNCH BREAK", "MORNING BREAK", "EVENING BREAK"]:
                    cell.fill = break_fill
                    cell.font = break_font
                elif cell_value and cell_value != "---":
                    if "Lab-" in str(cell_value):  # Lab session
                        cell.fill = lab_fill
                        cell.font = content_font
                    elif "LR-" in str(cell_value):  # Lecture session
                        cell.fill = lecture_fill
                        cell.font = content_font
                else:  # Empty slot
                    cell.fill = empty_fill
                    cell.font = content_font

    def export_teacher_timetables_to_excel(self, filename):
        """Export individual teacher timetables to the same Excel file as class timetables."""
        with pd.ExcelWriter(filename, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
            for teacher in self.teachers:
                # Build timetable grid: days x time slots
                data = []
                for time_slot in self.all_time_slots:
                    row = [time_slot]
                    for day in self.days:
                        # Find all assignments for this teacher at this slot
                        cell_entries = []
                        # Lectures
                        for class_name in self.all_classes:
                            lectures = self.class_timetables.get(class_name, {}).get('lectures', {})
                            entry = lectures.get(day, {}).get(time_slot)
                            if entry and entry.get('teacher') == teacher:
                                subj = entry['subject']
                                room = entry['room']
                                cell_entries.append(f"{subj} ({class_name}) {room}")
                        
                        # Labs - iterate through all classes
                        for class_name in self.all_classes:
                            lab_timetable = self.lab_schedule.get(class_name, None)
                            if lab_timetable:
                                labs = lab_timetable.get('labs', {}).get(day, {}).get(time_slot, {})
                                for batch, lab in labs.items():
                                    if isinstance(lab, dict) and lab.get('teacher') == teacher and not lab.get('continued', False):
                                        subj = lab['subject']
                                        lab_room = lab['lab']
                                        cell_entries.append(f"{subj} ({class_name}-{batch}) {lab_room}")
                                    elif isinstance(lab, dict) and lab.get('teacher') == teacher and lab.get('continued', False):
                                        # Show continued labs in the second hour
                                        subj = lab['subject']
                                        lab_room = lab['lab']
                                        cell_entries.append(f"{subj} ({class_name}-{batch}) {lab_room} (cont.)")
                        
                        if time_slot in self.break_slots:
                            if time_slot == self.lunch_break:
                                row.append("LUNCH BREAK")
                            elif time_slot == self.morning_break:
                                row.append("MORNING BREAK")
                            elif time_slot == self.evening_break:
                                row.append("EVENING BREAK")
                        elif cell_entries:
                            row.append("; ".join(cell_entries))
                        else:
                            row.append("---")
                    data.append(row)
                df = pd.DataFrame(data, columns=["Time Slot"] + self.days)
                df.to_excel(writer, sheet_name=teacher, index=False)
                # Style the worksheet
                workbook = writer.book
                worksheet = writer.sheets[teacher]
                self._style_timetable_worksheet(worksheet, teacher)

    def print_combined_timetable(self, class_name):
        """Print combined timetable with both lectures and labs"""
        if class_name not in self.class_timetables:
            print(f"No timetable available for class: {class_name}")
            return
        
        timetable = self.class_timetables[class_name]
        lab_timetable = self.lab_schedule[class_name] if class_name in self.lab_schedule else None
        
        print(f"\n{'='*120}")
        print(f"COMBINED TIMETABLE FOR {class_name} - {self._get_year_from_class(class_name)}")
        print(f"{'='*120}")
        
        time_header = "Time Slot"
        header = f"{time_header:<15} | "
        for day in self.days:
            header += f"{day:<20} | "
        print(header)
        print("-" * 120)
        
        for time_slot in self.all_time_slots:
            row = f"{time_slot:<15} | "
            for day in self.days:
                if time_slot in self.break_slots:
                    if time_slot == self.lunch_break:
                        row += f"{'LUNCH BREAK':<20} | "
                    elif time_slot == self.morning_break:
                        row += f"{'MORNING BREAK':<20} | "
                    elif time_slot == self.evening_break:
                        row += f"{'EVENING BREAK':<20} | "
                    continue
                
                # Check for lecture
                lecture = timetable["lectures"][day].get(time_slot)
                
                # Check for labs
                labs = []
                if lab_timetable:
                    for batch, lab in lab_timetable["labs"][day].get(time_slot, {}).items():
                        if lab and not lab.get("continued", False):
                            labs.append(f"{lab['subject']},{batch},{lab['lab']}")
                
                # Combine lecture and labs info
                if lecture:
                    cell = f"{lecture['subject']} ({lecture['teacher']}) ({lecture['room']})"
                    row += f"{cell:<20} | "
                elif labs:
                    # Format labs with teacher names
                    formatted_labs = []
                    for lab_info in labs:
                        # Extract subject, batch, lab from lab_info
                        parts = lab_info.split(',')
                        if len(parts) >= 3:
                            subject = parts[0]
                            batch = parts[1]
                            lab = parts[2]
                            # Get teacher for this subject
                            year = self._get_year_from_class(class_name)
                            teacher = self.teacher_assignments[year].get(subject, "Unknown")
                            formatted_labs.append(f"{subject} ({teacher}) {batch} {lab}")
                        else:
                            formatted_labs.append(lab_info)
                    cell = " && ".join(formatted_labs)
                    if len(cell) > 18:
                        cell = cell[:17] + "…"
                    row += f"{cell:<20} | "
                else:
                    row += f"{'---':<20} | "
            
            print(row)

def main():
    """Main function to run the lecture scheduler"""
    # Import the lab schedule from generator1
    from generator1 import main as run_lab_generator
    
    # First, run the lab generator to get the lab schedule
    print("Generating lab schedule first...")
    lab_generator = run_lab_generator()
    
    if not lab_generator:
        print("Failed to generate lab schedule. Cannot proceed with lecture scheduling.")
        return
    
    # Configuration for lecture scheduling
    years = ["Second Year", "Third Year"]
    classes_per_year = 1
    subjects_by_year = {
        "Second Year": ["ADE", "OS", "SDA", "PDS", "FLS", "UHV", "FLB"],
        "Third Year": ["DBMS", "AJP", "DC", "MNA", "BT"]
    }
    
    teacher_assignments = {
        "Third Year": {
            "DBMS": "Patil",
            "AJP": "Yadav",
            "DC": "Chimmana",
            "MNA": "Dolli",
            "BT": "Vidhate"
        },
        "Second Year": {
            "ADE": "Moon",
            "OS": "Chimmana",
            "SDA": "Nahatkar",
            "PDS": "Patil",
            "FLS": "Vidhate",
            "UHV": "Vidhate",
            "FLB": "Yadav"
        }
    }
    
    all_teachers = set()
    for year_map in teacher_assignments.values():
        for teacher in year_map.values():
            all_teachers.add(teacher)
    all_teachers = list(all_teachers)
    
    rooms = [str(506 + i) for i in range(2)]
    course_structure = {}
    for year, year_subjects in subjects_by_year.items():
        for subject in year_subjects:
            # Restore original lecture counts
            if year == "Third Year":
                if subject in ["DBMS", "DC", "BT", "MNA", "AJP"]:
                    lecture_count = 3  # Restored to 3
                else:
                    lecture_count = 2
            elif year == "Second Year":
                if subject in ["ADE", "PDS"]:
                    lecture_count = 3  # Restored to 3
                elif subject in ["UHV", "FLS", "OS", "FLB", "SDA"]:
                    lecture_count = 2
                else:
                    lecture_count = 2
            else:
                lecture_count = 2
                
            course_structure[subject] = {
                "lectures": lecture_count,
                "lecture_duration": THEORY_CLASS_DURATION
            }
    
    # Define allowed slots for Yadav and Vidhate
    allowed_slots_yadav_vidhate = [
        "10:30-11:30", "11:30-12:30", "1:15-2:15", "2:15-3:15", "3:30-4:30", "4:30-5:30"
    ]
    
    teacher_availability = {}
    for teacher in all_teachers:
        teacher_availability[teacher] = {day: [] for day in ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday"]}
    
    # Restrict Moon, Chimmana, Nahatkar, Patil, Dolli to 8:15-3:15 slots only
    for teacher in ["Moon", "Chimmana", "Nahatkar", "Patil", "Dolli"]:
        if teacher in teacher_availability:
            for day in teacher_availability[teacher]:
                teacher_availability[teacher][day] = [
                    "8:15-9:15", "9:15-10:15", "10:30-11:30", "11:30-12:30", 
                    "1:15-2:15", "2:15-3:15"
                ]
    
    # Restrict Yadav and Vidhate to 10:30-5:30 slots only
    for teacher in ["Yadav", "Vidhate"]:
        if teacher in teacher_availability:
            for day in teacher_availability[teacher]:
                teacher_availability[teacher][day] = [
                    "10:30-11:30", "11:30-12:30", "1:15-2:15", "2:15-3:15", 
                    "3:30-4:30", "4:30-5:30"
                ]
    
    department_data = {
        "years": years,
        "classes_per_year": classes_per_year,
        "teachers": all_teachers,
        "rooms": rooms,
        "subjects_by_year": subjects_by_year,
        "course_structure": course_structure,
        "teacher_assignments": teacher_assignments,
        "teacher_availability": teacher_availability
    }
    
    print("\nScheduling lectures around existing lab slots...")
    print(f"Years: {', '.join(years)}")
    print(f"Classes per year: {classes_per_year}")
    for year in years:
        print(f"\n{year} Lecture Subjects: {', '.join(subjects_by_year[year])}")
        print(f"{year} Lecture Teachers:")
        for subject, teacher in teacher_assignments[year].items():
            print(f"  - {subject}: {teacher}")
    print(f"\nLecture Rooms: {', '.join(rooms)}")
    print("=" * 50)
    
    # Create lecture scheduler with lab schedule
    scheduler = LectureScheduler(department_data, lab_generator.class_timetables)
    
    if scheduler.solve():
        print("\nLecture scheduling completed successfully!")
        
        # Print to console
        for class_name in scheduler.all_classes:
            scheduler.print_combined_timetable(class_name)
        
        # Export to Excel
        excel_filename = scheduler.export_combined_timetable_to_excel()
        print(f"\n📊 Excel file created: {excel_filename}")
        print("🎨 Features:")
        print("   • Color-coded cells (Green for labs, Yellow for lectures, Orange for breaks)")
        print("   • Professional formatting with borders and fonts")
        print("   • Auto-adjusted column widths")
        print("   • Timestamp in filename and worksheet")
        print("   • Separate sheets for each class")
        # Export teacher timetables
        scheduler.export_teacher_timetables_to_excel(excel_filename)
        print("   • Individual sheets for each teacher with their timetable")
    else:
        print("No feasible lecture schedule found.")

if __name__ == "__main__":
    main()
